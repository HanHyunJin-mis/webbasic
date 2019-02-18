# 파이썬 엑셀 처리
# pip3 install openpyxl - 엑셀 2007이상 버전에서 동작
# load_workbook - 저장된 엑셀 파일 불러오기
# create_sheet - 새로운 시트 만들기
# save - 엑셀파일 저장

from openpyxl import load_workbook as load 

DIR ='/Users/Administrator/Desktop/webbasic/python/rush.xlsx'

wb = load(DIR)

# ws = wb.create_sheet('test')
# ws['A1'] = "제목1"
# ws['B1'] = "제목2"

ws = wb['test']
a1 = ws['A2'].value
a2 = ws['B2'].value
print(a1, a2)

# wb.save(DIR)
wb.close()