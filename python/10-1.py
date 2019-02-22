from openpyxl import load_workbook as load 
DIR = '/Users/Administrator/Desktop/webbasic/python/travel.xlsx'

def save_excel(ws):
    idx = 65
    for i in range(1, 10):
        for j in range(1, 10):
            ws[chr(idx)+ str(j)] = "{} * {} = {}".format(i,j,i*j)
        idx += 1
    wb.save(DIR)

def get_excel(ws):
    idx = 64
    for i in range(1, 10):
        for j in range(1, 10):
            print(ws[chr(idx + j) + str(i)].value, end='')
        print()


wb = load(DIR)
ws = wb.create_sheet("99단")
try:
    save_excel(ws)
    get_excel(ws)
except Exception as e:
    print(e)
finally:
    wb.close()
